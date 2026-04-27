import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import os
import pyodbc
from django.conf import settings

class GetExcelInlaiBB:
    @staticmethod
    def create_excel(partno, effdat, path_file, slipno, barcode, ca, daylimt, soluong,
                     mesid, machno, pday, indat, classs, intime, pallet, loaikeo, description,masosx):
        try:
            # Xóa file nếu đã tồn tại
            if os.path.exists(path_file):
                os.remove(path_file)

            # Mô phỏng logic maso (thay bằng truy vấn cơ sở dữ liệu thực tế nếu cần)
            maso = masosx  # Giá trị mặc định

            # maso = get_maso(soluong, pday, partno, mesid, barcode)      // sau khi thêm số mẻ thì không cần sử dụng hàm auto nữa
           
            # Khởi tạo workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "1"  # Tên worksheet khớp với C#

            # Xác định loại thẻ
            tenloaithe = ""
            tentheloaikeoTQ = ""
            KVS = ""
            kichthuoc = ""
            if loaikeo == "RC":
                tenloaithe = "Thẻ biểu thị keo tinh luyện"
                tentheloaikeoTQ = "精煉膠標示卡"
                KVS = "KVS3J1C001.9 Rev.5"
                kichthuoc = "(180mm×130mm×0.08mm)"
            elif loaikeo in ["RD", "RR"]:
                tenloaithe = "Thẻ biểu thị keo xúc tiến"
                tentheloaikeoTQ = "加促膠標示卡"
                KVS = "KVS3J1C001.8 Rev.5"
                kichthuoc = "(180mm×130mm×0.08mm)"
            elif loaikeo == "RB":
                tenloaithe = "Thẻ biểu thị keo cán luyện"
                tentheloaikeoTQ = "混煉膠標示卡"
                KVS = "KVS3J1C001.7 Rev.5"
                kichthuoc = "(180mm×130mm×0.08mm)"

            # Ca làm việc
            Ca = "早 (Sáng)" if ca == "1" else "夜 (Đêm)" if ca == "2" else ""

            # Gán giá trị cho các ô
            ws["C2"] = " 建大橡膠（越南）有限公司"
            ws["C3"] = "Công ty Cao su Kenda(Việt Nam)"
            if description == "OEM":
                ws["M4"] = "QC-OEM"
            ws["E4"] = tentheloaikeoTQ
            ws["E5"] = tenloaithe
            ws["N5"] = pallet
            ws["C5"] = f"Machine {machno[2:8]}" if len(machno) >= 8 else f"Machine {machno}"
            ws["C6"] = f"日限 Thời hạn sử dụng: {daylimt}日 {daylimt}Ngày"
            ws["C7"] = "禁止雨淋，油污，置地，及粉水未乾"
            ws["C8"] = "Cấm ướt mưa, dính dầu, để lên đất, bột nước chưa khô"
            ws["C9"] = "生產日期 Ngày Tháng Sản Xuất"
            ws["C10"] = f"{indat} {intime}"
            ws["C11"] = "批號"
            ws["C12"] = "Số lô"
            ws["E11"] = slipno
            ws["I11"] = "重量"
            ws["I12"] = "Trọng lượng"
            ws["L11"] = f"{soluong}KG"
            ws["C13"] = "有效日"
            ws["C14"] = "Ngày hiệu lực"
            ws["E13"] = f"{effdat} {intime}"
            ws["I13"] = "班別"
            ws["I14"] = "Ca"
            ws["L13"] = Ca
            ws["C15"] = "規格"
            ws["C16"] = "Quy Cách"
            
            
            ws["D15"] = partno
            ws["G15"] = "編號順序"
            ws["G17"] = "Thứ tự mã số"
            ws["I15"] = maso
            ws["K15"] = "判 定"
            ws["K17"] = "Phán định"
            ws["M15"] = ""  # Để trống theo C#
            ws["C17"] = "委託"
            ws["C18"] = "Ủy thác"
            ws["C20"] = KVS
            ws["K20"] = f"*{barcode}*"
            ws["C21"] = kichthuoc
            ws["K21"] = f"*{barcode}*"

            # Gộp ô
            merges = [
                "M4:O4", "C5:D5", "C2:O2", "C3:O3", "E4:L4", "E5:L5", "N5:O5", "C6:O6",
                "C7:O7", "C8:O8", "C9:O9", "C10:O10", "C11:D11", "C12:D12", "E11:H12",
                "I11:K11", "I12:K12", "L11:O12", "C13:D13", "C14:D14", "E13:H14",
                "I13:K13", "I14:K14", "L13:O14", "D15:F16", "G15:H16", "G17:H18",
                "I15:J18", "K15:L16", "K17:L18", "M15:O18", "D17:F18", "C20:G20",
                "K20:O20", "C21:G21", "K21:O21"
            ]
            for m in merges:
                ws.merge_cells(m)

            # Thiết lập kiểu dáng
            font_default = Font(name="Arial", size=13, color="000000")
            font_bold = Font(name="Arial", size=13, bold=True, color="000000")
            font_barcode = Font(name="Code39AzaleaWide2", size=28)
            font_small = Font(name="Arial", size=10)
            font_medium = Font(name="Arial", size=12)
            font_large = Font(name="Arial", size=16)
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style="thin"), bottom=Side(style="thin"))
            medium_border = Border(left=Side(style="medium"), right=Side(style="medium"),
                                   top=Side(style="medium"), bottom=Side(style="medium"))

            # Áp dụng kiểu mặc định cho vùng B2:P22
            for row in ws["B2:P22"]:
                for cell in row:
                    cell.font = font_default
                    cell.alignment = align_center
                    cell.fill = fill_white
                    cell.border = thin_border

            remove_borders(ws["B2"], keep=("left","top"))
            remove_borders(ws["P2"], keep=("right","top"))
            
            for row in ws["C2:O2"]:
                for cell in row:
                    remove_borders(cell, keep=("top"))
            remove_borders(ws["B3"], keep=("left"))
          
            for row in ws["C3:O3"]:
                for cell in row:
                    remove_borders(cell, keep=())
            
            remove_borders(ws["P3"], keep=("right"))
            remove_borders(ws["B4"], keep=("left"))
            remove_borders(ws["P4"], keep=("right"))

           
            for row in ws["C4:O4"]:
                for cell in row:
                    remove_borders(cell, keep=())
            for row in ws["C5:O5"]:
                for cell in row:
                    remove_borders(cell, keep=("bottom"))    

            remove_borders(ws["B5"], keep=("left"))
            remove_borders(ws["P5"], keep=("right"))    

            for row in ws["B6:B18"]:
                for cell in row:
                    remove_borders(cell, keep=("left","right"))
            for row in ws["P6:P18"]:
                for cell in row:
                    remove_borders(cell, keep=("left","right"))

            for row in ws["C19:O19"]:
                for cell in row:
                    remove_borders(cell, keep=("top"))
            remove_borders(ws["B19"], keep=("left"))
            remove_borders(ws["P19"], keep=("right")) 

            for row in ws["C20:O20"]:
                for cell in row:
                    remove_borders(cell, keep=())

            remove_borders(ws["B20"], keep=("left"))
            remove_borders(ws["P20"], keep=("right"))
            for row in ws["C21:O21"]:
                for cell in row:
                    remove_borders(cell, keep=())
            remove_borders(ws["B21"], keep=("left"))
            remove_borders(ws["P21"], keep=("right"))
            for row in ws["C22:O22"]:
                for cell in row:
                    remove_borders(cell, keep=("bottom"))
            remove_borders(ws["B22"], keep=("left","bottom"))
            remove_borders(ws["P22"], keep=("right","bottom"))
            
            




            # Kiểu dáng cụ thể cho từng ô
            ws["M4"].font = font_large
            ws["M4"].font = font_bold
            ws["E4"].font = font_bold
            ws["E5"].font = font_bold
            ws["K20"].font = font_barcode
            ws["K21"].font = font_default
            ws["C16"].font = font_medium
            ws["C5"].font = font_small
            ws["C2"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
            ws["C7"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
            ws["C8"].alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            ws["G15"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
            ws["K15"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
            ws["G17"].alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            ws["K17"].alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            ws["E14"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Sửa lỗi viền: Lặp qua từng ô thay vì hàng
            for row in ws["C7:O7"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=cell.border.top, bottom=Side(style=None))
            for row in ws["C8:O8"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=Side(style=None), bottom=cell.border.bottom)
            for row in ws["C11:K11"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=cell.border.top, bottom=Side(style=None))
            for row in ws["C12:K12"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=Side(style=None), bottom=cell.border.bottom)
            for row in ws["C13:D13"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=cell.border.top, bottom=Side(style=None))
            for row in ws["C14:D14"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=Side(style=None), bottom=cell.border.bottom)
            for row in ws["I13:K13"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=cell.border.top, bottom=Side(style=None))
            for row in ws["I14:K14"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=Side(style=None), bottom=cell.border.bottom)
            for row in ws["C15:C15"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=cell.border.top, bottom=Side(style=None))
            # for row in ws["C16:C16"]:
            #     for cell in row:
            #         cell.border = Border(left=cell.border.left, right=cell.border.right,
            #                             top=Side(style=None), bottom=cell.border.bottom)
            for row in ws["C16:C16"]:
                for cell in row:
                    # Giữ border gốc (trừ top)
                    cell.border = Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=Side(style=None),
                        bottom=cell.border.bottom
                    )
                    # Tắt wrap text, vẫn giữ nguyên canh lề cũ
                    cell.alignment = Alignment(
                        wrapText=False,
                        horizontal=cell.alignment.horizontal if cell.alignment else None,
                        vertical=cell.alignment.vertical if cell.alignment else None
                    )
            for row in ws["C17:C17"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=cell.border.top, bottom=Side(style=None))
            for row in ws["C18:C18"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=Side(style=None), bottom=cell.border.bottom)
            for row in ws["G16:H16"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=cell.border.top, bottom=Side(style=None))
            for row in ws["G17:H17"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=Side(style=None), bottom=cell.border.bottom)
            for row in ws["K16:L16"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=cell.border.top, bottom=Side(style=None))
            for row in ws["K17:L17"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=Side(style=None), bottom=cell.border.bottom)

            # Áp dụng viền dày cho vùng B2:P22
            for row in ws["B2:P2"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=Side(style="medium"), bottom=cell.border.bottom)
            for row in ws["B22:P22"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=cell.border.right,
                                        top=cell.border.top, bottom=Side(style="medium"))
            for row in ws["B2:B22"]:
                for cell in row:
                    cell.border = Border(left=Side(style="medium"), right=cell.border.right,
                                        top=cell.border.top, bottom=cell.border.bottom)
            for row in ws["P2:P22"]:
                for cell in row:
                    cell.border = Border(left=cell.border.left, right=Side(style="medium"),
                                        top=cell.border.top, bottom=cell.border.bottom)

            # Thiết lập chiều cao hàng
            ws.row_dimensions[1].height = 5
            ws.row_dimensions[2].height = 20.5
            ws.row_dimensions[7].height = 22.5
            ws.row_dimensions[8].height = 22.5
            ws.row_dimensions[19].height = 5
            ws.row_dimensions[20].height = 30
            ws.row_dimensions[22].height = 5.5

            # Thiết lập chiều rộng cột
            ws.column_dimensions["A"].width = 2
            ws.column_dimensions["B"].width = 2
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 5
            ws.column_dimensions["E"].width = 5
            ws.column_dimensions["F"].width = 5
            ws.column_dimensions["G"].width = 8
            ws.column_dimensions["H"].width = 8
            ws.column_dimensions["I"].width = 5
            ws.column_dimensions["J"].width = 5
            ws.column_dimensions["K"].width = 5
            ws.column_dimensions["L"].width = 7
            ws.column_dimensions["M"].width = 5
            ws.column_dimensions["N"].width = 5
            ws.column_dimensions["O"].width = 5
            ws.column_dimensions["P"].width = 2

            # Thiết lập lề trang và cài đặt in
            ws.page_margins = PageMargins(top=0.2, left=0.8, right=0.5)
            ws.page_setup.paperSize = 11  # A5
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1

            # Thiết lập lề trang và cài đặt in



            # Lưu file
            wb.save(path_file)
            return path_file  # Trả về giá trị rỗng theo C#
        except Exception as e:
            return f"Lỗi file Excel: {str(e)}"
        
        
def remove_borders(cell, keep=("left","top")):
    cell.border = Border(
        left=cell.border.left if "left" in keep else Side(style=None),
        right=cell.border.right if "right" in keep else Side(style=None),
        top=cell.border.top if "top" in keep else Side(style=None),
        bottom=cell.border.bottom if "bottom" in keep else Side(style=None),
    )
def remove_borders_range(ws, cell_range, keep=()):
    """Xoá border trong cả vùng (dù merge hay không)."""
    for row in ws[cell_range]:
        for cell in row:
            remove_borders(cell, keep=keep)

def get_maso(Soluong, pday, partno, mesid, barcode):
    maso = ""

    if int(Soluong) <= 380:
        maso = "1"
    else:
        sql = f"""
            SELECT barcode 
            FROM prdebe
            WHERE factory='V' 
              AND prodat='{pday}'
              AND partno='{partno}'
              AND mesid='{mesid}'
              AND barcode <= '{barcode}'
            ORDER BY intime
        """
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(sql)
        rows = cursor.fetchall()
        conn.close()

        a, b = 1, 2
        if rows:
            for i in range(len(rows)):
                if i > 0:
                    a += 2
                    b = a + 1
            maso = f"{a}-{b}"

    return maso

def get_connection():
    db_config = settings.DATABASES['default']
    conn = pyodbc.connect(
        f"DRIVER={db_config['OPTIONS']['driver']};"
        f"SERVER={db_config['HOST']};"
        f"DATABASE={db_config['NAME']};"
        f"UID={db_config['USER']};"
        f"PWD={db_config['PASSWORD']};"
        f"{db_config['OPTIONS']['extra_params']}"
    )
    return conn


